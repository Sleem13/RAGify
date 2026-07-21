from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
import os
import io
import csv
import json
import secrets
import logging
from typing import Optional

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

from services.document_processor import document_processor
from services.vector_db import vector_db
from services.llm_manager import llm_manager
from services.data_analyzer import data_analyzer
from services.retrieval import build_context, build_grounded_prompt, normalize_history

app = FastAPI(
    title="RAGify API",
    description="AI-powered document analysis and retrieval system. Upload files, ask questions, get insights.",
    version="5.0.0",
    docs_url="/docs",
    redoc_url="/redoc",
)

# ─── CORS ─────────────────────────────────────────────────────────────────────
configured_origins = [
    origin.strip()
    for origin in os.getenv("FRONTEND_URL", "*").split(",")
    if origin.strip()
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=configured_origins or ["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_EXTENSIONS = {".xlsx", ".xls", ".csv", ".json"}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB hard limit

# ─── File Registry — persisted to disk ───────────────────────────────────────
# Survives backend restarts unlike a plain in-memory dict.
_api_keys: set[str] = set()

REGISTRY_PATH = os.path.join(os.path.dirname(__file__), "data", "files_registry.json")


def _load_registry() -> dict:
    """Load the file registry from disk on startup."""
    os.makedirs(os.path.dirname(REGISTRY_PATH), exist_ok=True)
    if os.path.exists(REGISTRY_PATH):
        try:
            with open(REGISTRY_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as exc:
            logger.warning(f"Could not read registry: {exc}")
    return {}


def _save_registry(registry: dict):
    """Persist the file registry to disk."""
    os.makedirs(os.path.dirname(REGISTRY_PATH), exist_ok=True)
    try:
        with open(REGISTRY_PATH, "w", encoding="utf-8") as f:
            json.dump(registry, f, ensure_ascii=False, indent=2)
    except Exception as exc:
        logger.error(f"Could not save registry: {exc}")


# Load registry on startup
_uploaded_files: dict = _load_registry()
if _uploaded_files and not vector_db.has_index:
    logger.warning("Ignoring stale file registry because the FAISS index is missing.")
    _uploaded_files = {}
logger.info(f"Loaded file registry: {len(_uploaded_files)} file(s) found.")


# ─── API Key helpers ──────────────────────────────────────────────────────────

def _verify_api_key(x_api_key: str | None) -> bool:
    """If no keys have been generated yet, allow all traffic (open mode)."""
    if not _api_keys:
        return True
    return x_api_key in _api_keys


# ═════════════════════════════════════════════════════════════════════════════
#  HEALTH CHECK
# ═════════════════════════════════════════════════════════════════════════════

@app.get("/", tags=["System"])
def health_check():
    return {
        "status": "healthy",
        "app": "RAGify API",
        "version": "5.0.0",
        "loaded_models": llm_manager.available_models,
        "indexed_files": len(_uploaded_files),
        "docs": "/docs",
    }


# ═════════════════════════════════════════════════════════════════════════════
#  API KEY MANAGEMENT
# ═════════════════════════════════════════════════════════════════════════════

@app.post("/generate-api-key", tags=["API Key"])
def generate_api_key():
    new_key = "ragify-" + secrets.token_urlsafe(32)
    _api_keys.add(new_key)
    return {
        "api_key": new_key,
        "message": "Keep this key safe. Use it in the 'X-Api-Key' header for all API requests.",
    }


@app.get("/list-api-keys", tags=["API Key"])
def list_api_keys():
    masked = [k[:12] + "..." + k[-4:] for k in _api_keys]
    return {"active_keys": masked, "total": len(_api_keys)}


@app.delete("/revoke-api-key", tags=["API Key"])
def revoke_api_key(api_key: str):
    if api_key in _api_keys:
        _api_keys.discard(api_key)
        return {"message": "API key revoked successfully."}
    raise HTTPException(status_code=404, detail="API key not found.")


# ═════════════════════════════════════════════════════════════════════════════
#  FILE MANAGEMENT
# ═════════════════════════════════════════════════════════════════════════════

@app.get("/files", tags=["Documents"])
def list_files():
    """List all currently indexed files."""
    return {"files": _uploaded_files, "total": len(_uploaded_files)}


@app.delete("/files/{filename}", tags=["Documents"])
def delete_file(filename: str, x_api_key: str | None = Header(default=None)):
    """
    Remove a specific file from the registry AND from the FAISS vector database.
    After this call the AI will no longer have access to that file's content.
    """
    if not _verify_api_key(x_api_key):
        raise HTTPException(status_code=401, detail="Invalid or missing API key.")
    if filename not in _uploaded_files:
        raise HTTPException(status_code=404, detail=f"File '{filename}' not found in registry.")

    # Remove from FAISS (surgical delete by source metadata)
    deleted_from_faiss = vector_db.delete_by_filename(filename)
    if deleted_from_faiss:
        logger.info(f"'{filename}' vectors removed from FAISS.")
    else:
        logger.warning(f"'{filename}' had no vectors in FAISS (may have already been removed).")

    # Remove from registry and persist
    del _uploaded_files[filename]
    _save_registry(_uploaded_files)

    return {
        "message": f"✅ '{filename}' removed from registry and vector database.",
        "faiss_deleted": deleted_from_faiss,
    }


# ═════════════════════════════════════════════════════════════════════════════
#  FILE UPLOAD & PROCESSING
# ═════════════════════════════════════════════════════════════════════════════

@app.post("/upload", tags=["Documents"])
async def upload_file(
    file: UploadFile = File(...),
    x_api_key: str | None = Header(default=None),
):
    """
    Upload and process any supported file.
    - PDF, DOCX, TXT, PPTX, Images → RAG pipeline (indexed in FAISS for Q&A)
    - XLSX, XLS, CSV → Dual pipeline: analyzed with Pandas AND indexed in FAISS for Q&A
    
    Max file size: 50 MB
    """
    if not _verify_api_key(x_api_key):
        raise HTTPException(status_code=401, detail="Invalid or missing API key.")

    filename = file.filename or "unknown"
    ext = os.path.splitext(filename)[1].lower()
    contents = await file.read()

    if not contents:
        raise HTTPException(status_code=400, detail="The uploaded file is empty.")

    # ── File size guard ───────────────────────────────────────────────────────
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"File too large ({len(contents) / 1024 / 1024:.1f} MB). Maximum allowed size is 50 MB."
        )

    # ── Excel / CSV / JSON files → DUAL pipeline (analytics + RAG) ───────────
    if ext in EXCEL_EXTENSIONS:
        # 1. Analytics pipeline
        try:
            analysis = await data_analyzer.analyze_excel(contents, filename)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Analysis error: {exc}")

        # 2. Also index the data into FAISS for Q&A
        chunks = []
        try:
            chunks = await document_processor.process_excel_for_rag(contents, filename)
            if not chunks:
                raise ValueError("The spreadsheet did not contain indexable data.")
            vector_db.replace_file_chunks(chunks, filename=filename)
            logger.info(f"Excel '{filename}' indexed in FAISS with {len(chunks)} chunks.")
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"Indexing error: {exc}") from exc

        _uploaded_files[filename] = {
            "type": "excel",
            "chunks": len(chunks),
            "charts": len(analysis.get("charts", [])),
        }
        _save_registry(_uploaded_files)

        return {
            "filename": filename,
            "status": "analyzed",
            "type": "excel",
            "analysis": analysis,
            "message": f"✅ '{filename}' analyzed and indexed. You can now ask questions about this data AND view the dashboard.",
        }

    # ── Document files (RAG pipeline) ────────────────────────────────────────
    if document_processor.is_supported(filename):
        try:
            chunks = await document_processor.process_file(contents, filename)
        except ValueError as exc:
            raise HTTPException(status_code=415, detail=str(exc))
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Processing error: {exc}")

        if not chunks:
            raise HTTPException(
                status_code=422,
                detail="No readable content was found in the uploaded document.",
            )
        vector_db.replace_file_chunks(chunks, filename=filename)
        _uploaded_files[filename] = {"type": "document", "chunks": len(chunks)}
        _save_registry(_uploaded_files)

        # Generate a smart summary after indexing
        prompt = (
            f"A document named '{filename}' was just uploaded and indexed with {len(chunks)} text chunks.\n"
            "In 2-3 sentences, acknowledge this professionally and mention you are ready for questions.\n"
            "MUST respond in English."
        )
        try:
            summary = await llm_manager.generate_response(prompt)
        except Exception:
            summary = f"✅ '{filename}' has been indexed ({len(chunks)} chunks). Ask me anything about it!"

        return {
            "filename": filename,
            "status": "processed",
            "type": "document",
            "chunks": len(chunks),
            "message": summary,
        }

    raise HTTPException(
        status_code=415,
        detail=f"Unsupported file type '{ext}'. Supported: PDF, DOCX, TXT, PPTX, PNG, JPG, JPEG, XLSX, XLS, CSV.",
    )


# ═════════════════════════════════════════════════════════════════════════════
#  CHAT (RAG + Conversation Memory + Agentic Actions)
# ═════════════════════════════════════════════════════════════════════════════

@app.post("/chat", tags=["Chat"])
async def chat(
    message: str = Form(...),
    history: str = Form(default="[]"),
    x_api_key: str | None = Header(default=None),
):
    """
    RAG Chat endpoint with conversation memory.
    Accepts `history` as a JSON array of [{role, content}] for context continuity.
    """
    if not _verify_api_key(x_api_key):
        raise HTTPException(status_code=401, detail="Invalid or missing API key.")

    if not message.strip():
        raise HTTPException(status_code=400, detail="Message cannot be empty.")

    # ── Parse conversation history ─────────────────────────────────────────────
    try:
        chat_history = normalize_history(json.loads(history))
    except (TypeError, json.JSONDecodeError):
        chat_history = []

    # ── Retrieve context from FAISS Vector DB ─────────────────────────────────
    try:
        matches = vector_db.hybrid_search(message, top_k=10)
    except Exception as exc:
        logger.warning("Hybrid retrieval failed: %s", exc)
        matches = []

    context, sources = build_context(matches)
    if not context:
        response_text = (
            "I could not find this information in the uploaded files."
            if _uploaded_files
            else "Upload a document or spreadsheet first, then ask a question about it."
        )
        return {
            "response": response_text,
            "context_found": False,
            "action": None,
            "sources": [],
        }

    # ── Build conversation history string ──────────────────────────────────────
    # ── Build prompt ──────────────────────────────────────────────────────────
    # ── Generate response ─────────────────────────────────────────────────────
    prompt = build_grounded_prompt(
        message,
        context,
        chat_history,
        _uploaded_files.keys(),
    )

    try:
        response_text = await llm_manager.generate_response(prompt)

        # Intercept Agent Actions
        action = None
        if "[ACTION: GENERATE_DASHBOARD]" in response_text:
            action = "GENERATE_DASHBOARD"
            response_text = response_text.replace("[ACTION: GENERATE_DASHBOARD]", "").strip()
            if not response_text:
                response_text = "I have analyzed the data and prepared your dashboard. Click the button below to open it."

        return {
            "response": response_text,
            "context_found": True,
            "action": action,
            "sources": sources,
        }
    except RuntimeError as exc:
        return {
            "response": "⚠️ All AI providers are currently unavailable. Please try again in a moment.",
            "context_found": False,
            "sources": sources,
            "error": str(exc),
        }


# ═════════════════════════════════════════════════════════════════════════════
#  EXPORT ANALYTICS DATA
# ═════════════════════════════════════════════════════════════════════════════

@app.post("/export", tags=["Export"])
async def export_data(
    format: str = Form("json"),
    data: str = Form(...),
    x_api_key: str | None = Header(default=None),
):
    try:
        parsed = json.loads(data)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON data provided.")

    fmt = format.lower().strip()

    if fmt == "json":
        content = json.dumps(parsed, indent=2, ensure_ascii=False)
        return StreamingResponse(
            io.BytesIO(content.encode("utf-8")),
            media_type="application/json",
            headers={"Content-Disposition": "attachment; filename=ragify_export.json"},
        )

    if fmt == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        summary = parsed.get("summary", {})
        columns = summary.get("columns", [])
        writer.writerow(["RAGify Analytics Export"])
        writer.writerow([])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Rows", summary.get("rows_count", "N/A")])
        writer.writerow(["Total Columns", len(columns)])
        writer.writerow(["Insights", parsed.get("insights", "")])
        writer.writerow([])
        writer.writerow(["Detected Columns"] + columns)
        chart = parsed.get("chart_data")
        if chart and chart.get("labels") and chart.get("datasets"):
            writer.writerow([])
            writer.writerow(["Chart Data"])
            writer.writerow(["Label", chart["datasets"][0].get("label", "Value")])
            for label, val in zip(chart["labels"], chart["datasets"][0]["data"]):
                writer.writerow([label, val])
        content = output.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=ragify_export.csv"},
        )

    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RAGify Analytics"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="6366F1")
        summary = parsed.get("summary", {})
        columns = summary.get("columns", [])
        ws.append(["RAGify Analytics Export"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        ws.append(["Metric", "Value"])
        for cell in ws[3]:
            cell.font = header_font
            cell.fill = header_fill
        ws.append(["Total Rows", summary.get("rows_count", "N/A")])
        ws.append(["Total Columns", len(columns)])
        ws.append(["Insights", parsed.get("insights", "")])
        ws.append([])
        ws.append(["Detected Columns"])
        ws.append(columns)
        chart = parsed.get("chart_data")
        if chart and chart.get("labels") and chart.get("datasets"):
            ws.append([])
            ws.append(["Chart Data"])
            ws.append(["Label", chart["datasets"][0].get("label", "Value")])
            for label, val in zip(chart["labels"], chart["datasets"][0]["data"]):
                ws.append([label, val])
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=ragify_export.xlsx"},
        )

    if fmt == "pdf":
        try:
            from fpdf import FPDF
        except ImportError:
            raise HTTPException(status_code=500, detail="fpdf2 not installed.")
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", style="B", size=16)
        pdf.cell(200, 10, txt="RAGify Analytics Report", ln=1, align="C")
        pdf.ln(8)
        summary = parsed.get("summary", {})
        columns = summary.get("columns", [])
        pdf.set_font("Helvetica", style="B", size=12)
        pdf.cell(200, 10, txt="Summary Metrics", ln=1)
        pdf.set_font("Helvetica", size=11)
        pdf.cell(200, 8, txt=f"Total Rows: {summary.get('rows_count', 'N/A')}", ln=1)
        pdf.cell(200, 8, txt=f"Total Columns: {len(columns)}", ln=1)
        pdf.ln(5)
        pdf.set_font("Helvetica", style="B", size=12)
        pdf.cell(200, 10, txt="AI Insights", ln=1)
        pdf.set_font("Helvetica", size=11)
        pdf.multi_cell(0, 8, txt=parsed.get("insights", ""))
        pdf.ln(5)
        pdf.set_font("Helvetica", style="B", size=12)
        pdf.cell(200, 10, txt="Columns Detected", ln=1)
        pdf.set_font("Helvetica", size=11)
        pdf.multi_cell(0, 8, txt=", ".join(columns))
        content = pdf.output()
        buffer = io.BytesIO(content)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=ragify_export.pdf"},
        )

    raise HTTPException(status_code=400, detail=f"Unknown format '{fmt}'. Use: json, csv, xlsx, pdf.")


# ═════════════════════════════════════════════════════════════════════════════
#  RESET KNOWLEDGE BASE
# ═════════════════════════════════════════════════════════════════════════════

@app.delete("/reset", tags=["System"])
def reset_knowledge_base(x_api_key: str | None = Header(default=None)):
    """Wipe the entire FAISS vector database and file registry."""
    if not _verify_api_key(x_api_key):
        raise HTTPException(status_code=401, detail="Invalid or missing API key.")
    vector_db.reset()

    # Clear and persist the registry
    _uploaded_files.clear()
    _save_registry(_uploaded_files)

    logger.info("FAISS vector database and file registry wiped.")
    return {"message": "✅ Knowledge base reset. All documents have been removed."}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=9999, reload=True)
