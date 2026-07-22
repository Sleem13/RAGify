"""Serialize analytics results into downloadable report formats."""

from __future__ import annotations

import csv
import io
import json
from typing import Any


ExportArtifact = tuple[bytes, str, str]


def export_analytics(format_name: str, data: dict[str, Any]) -> ExportArtifact:
    exporters = {
        "json": _json_export,
        "csv": _csv_export,
        "xlsx": _xlsx_export,
        "pdf": _pdf_export,
    }
    normalized_format = format_name.lower().strip()
    exporter = exporters.get(normalized_format)
    if exporter is None:
        raise ValueError(
            f"Unknown format '{normalized_format}'. Use: {', '.join(exporters)}."
        )
    return exporter(data)


def _json_export(data: dict[str, Any]) -> ExportArtifact:
    content = json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8")
    return content, "application/json", "ragify_export.json"


def _csv_export(data: dict[str, Any]) -> ExportArtifact:
    output = io.StringIO()
    writer = csv.writer(output)
    summary = data.get("summary", {})
    columns = summary.get("columns", [])
    writer.writerow(["RAGify Analytics Export"])
    writer.writerow([])
    writer.writerow(["Metric", "Value"])
    writer.writerow(["Total Rows", summary.get("rows_count", "N/A")])
    writer.writerow(["Total Columns", len(columns)])
    writer.writerow(["Insights", data.get("insights", "")])
    writer.writerow([])
    writer.writerow(["Detected Columns", *columns])
    _write_chart_rows(writer, data.get("chart_data"))
    return output.getvalue().encode("utf-8-sig"), "text/csv", "ragify_export.csv"


def _write_chart_rows(writer: Any, chart: Any) -> None:
    if not isinstance(chart, dict) or not chart.get("labels") or not chart.get("datasets"):
        return
    dataset = chart["datasets"][0]
    writer.writerow([])
    writer.writerow(["Chart Data"])
    writer.writerow(["Label", dataset.get("label", "Value")])
    for label, value in zip(chart["labels"], dataset.get("data", [])):
        writer.writerow([label, value])


def _xlsx_export(data: dict[str, Any]) -> ExportArtifact:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is not installed.") from exc

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "RAGify Analytics"
    summary = data.get("summary", {})
    columns = summary.get("columns", [])

    worksheet.append(["RAGify Analytics Export"])
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet.append([])
    worksheet.append(["Metric", "Value"])
    for cell in worksheet[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="6366F1")
    worksheet.append(["Total Rows", summary.get("rows_count", "N/A")])
    worksheet.append(["Total Columns", len(columns)])
    worksheet.append(["Insights", data.get("insights", "")])
    worksheet.append([])
    worksheet.append(["Detected Columns"])
    worksheet.append(columns)

    chart = data.get("chart_data")
    if isinstance(chart, dict) and chart.get("labels") and chart.get("datasets"):
        dataset = chart["datasets"][0]
        worksheet.append([])
        worksheet.append(["Chart Data"])
        worksheet.append(["Label", dataset.get("label", "Value")])
        for label, value in zip(chart["labels"], dataset.get("data", [])):
            worksheet.append([label, value])

    for column in worksheet.columns:
        max_length = max((len(str(cell.value or "")) for cell in column), default=10)
        worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 4, 50)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return (
        buffer.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "ragify_export.xlsx",
    )


def _pdf_export(data: dict[str, Any]) -> ExportArtifact:
    try:
        from fpdf import FPDF
    except ImportError as exc:
        raise RuntimeError("fpdf2 is not installed.") from exc

    summary = data.get("summary", {})
    columns = summary.get("columns", [])
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", style="B", size=16)
    pdf.cell(200, 10, text="RAGify Analytics Report", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(8)
    pdf.set_font("Helvetica", style="B", size=12)
    pdf.cell(200, 10, text="Summary Metrics", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", size=11)
    pdf.cell(
        200,
        8,
        text=f"Total Rows: {summary.get('rows_count', 'N/A')}",
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.cell(
        200,
        8,
        text=f"Total Columns: {len(columns)}",
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.ln(5)
    pdf.set_font("Helvetica", style="B", size=12)
    pdf.cell(200, 10, text="AI Insights", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", size=11)
    pdf.multi_cell(0, 8, text=str(data.get("insights", "")))
    pdf.ln(5)
    pdf.set_font("Helvetica", style="B", size=12)
    pdf.cell(200, 10, text="Columns Detected", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", size=11)
    pdf.multi_cell(0, 8, text=", ".join(map(str, columns)))
    return bytes(pdf.output()), "application/pdf", "ragify_export.pdf"
